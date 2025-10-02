from flask import Flask, render_template, request, redirect, url_for, flash, make_response, session
import secrets
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from functools import wraps
import mysql.connector
from mysql.connector import pooling
import logging
import os
from dotenv import load_dotenv

# Environment variables yükle
load_dotenv()

app = Flask(__name__)

# Environment variables'dan yapılandırma
app.secret_key = os.getenv('FLASK_SECRET_KEY', secrets.token_hex(16))

# Logging yapılandırması
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Memory-based rezervasyonlar (geçici - MySQL'e aktarılacak)
reservations = []

# MySQL Bağlantı Ayarları - Environment Variables'dan
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'database': os.getenv('DB_DATABASE', 'rezervasyon_sistemi'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', 'Password2025!'),
    'port': int(os.getenv('DB_PORT', 3306)),
    'charset': 'utf8mb4',
    'collation': 'utf8mb4_unicode_ci',
    'autocommit': True,
    'pool_name': 'mypool',
    'pool_size': 5,
    'pool_reset_session': True
}

# Bağlantı havuzu oluştur
try:
    connection_pool = mysql.connector.pooling.MySQLConnectionPool(**DB_CONFIG)
    logger.info("MySQL bağlantı havuzu oluşturuldu")
except mysql.connector.Error as err:
    logger.error(f"MySQL bağlantı hatası: {err}")
    connection_pool = None

def get_db_connection():
    """Veritabanı bağlantısı al"""
    try:
        if connection_pool:
            return connection_pool.get_connection()
        return None
    except mysql.connector.Error as err:
        logger.error(f"Bağlantı alma hatası: {err}")
        return None

def init_database():
    """Veritabanı ve tabloları oluştur"""
    try:
        # İlk olarak veritabanını oluştur
        temp_config = DB_CONFIG.copy()
        temp_config.pop('database', None)  # Database ismini kaldır
        
        temp_conn = mysql.connector.connect(**temp_config)
        cursor = temp_conn.cursor()
        
        # Veritabanını oluştur
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_CONFIG['database']} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        cursor.close()
        temp_conn.close()
        
        # Şimdi asıl veritabanına bağlan
        conn = get_db_connection()
        if not conn:
            return False
            
        cursor = conn.cursor()
        
        # Rezervasyonlar tablosu
        reservations_table = """
        CREATE TABLE IF NOT EXISTS reservations (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name_surname VARCHAR(255) NOT NULL,
            center VARCHAR(255) NOT NULL,
            venue VARCHAR(100) DEFAULT 'Tiyatro Salonu',
            date DATE NOT NULL,
            time VARCHAR(20) NOT NULL,
            description TEXT,
            status ENUM('onay', 'bekle', 'iptal') DEFAULT 'bekle',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            created_by VARCHAR(100),
            INDEX idx_center_date (center, date),
            INDEX idx_status (status),
            INDEX idx_created_at (created_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
        
        # Kullanıcılar tablosu
        users_table = """
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(100) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            role VARCHAR(50) DEFAULT 'user',
            permissions JSON,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_username (username)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
        
        cursor.execute(reservations_table)
        cursor.execute(users_table)
        
        # Varsayılan kullanıcıları ekle
        insert_default_users(cursor)
        
        conn.commit()
        cursor.close()
        conn.close()
        
        logger.info("Veritabanı tabloları başarıyla oluşturuldu")
        return True
        
    except mysql.connector.Error as err:
        logger.error(f"Veritabanı oluşturma hatası: {err}")
        return False

def insert_default_users(cursor):
    """Varsayılan kullanıcıları ekle"""
    default_users = [
        ('admin', 'admin123', 'admin', '["view_reservations", "edit_reservations", "view_availability", "manage_users"]'),
        ('editor1', 'editor123', 'editor', '["edit_reservations"]'),
        ('viewer1', 'viewer123', 'viewer', '["view_reservations"]'),
        ('scheduler1', 'scheduler123', 'scheduler', '["view_availability"]'),
        ('user1', 'user123', 'user', '[]')
    ]
    
    for username, password, role, permissions in default_users:
        try:
            cursor.execute("""
                INSERT IGNORE INTO users (username, password, role, permissions)
                VALUES (%s, %s, %s, %s)
            """, (username, password, role, permissions))
        except mysql.connector.Error as err:
            logger.error(f"Kullanıcı ekleme hatası ({username}): {err}")

def load_users_from_db():
    """Kullanıcıları veritabanından yükle"""
    try:
        conn = get_db_connection()
        if not conn:
            return {}
            
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT username, password, role, permissions FROM users")
        db_users = cursor.fetchall()
        
        users_dict = {}
        for user in db_users:
            import json
            permissions = json.loads(user['permissions']) if user['permissions'] else []
            users_dict[user['username']] = {
                'password': user['password'],
                'role': user['role'],
                'permissions': permissions
            }
        
        cursor.close()
        conn.close()
        return users_dict
        
    except mysql.connector.Error as err:
        logger.error(f"Kullanıcı yükleme hatası: {err}")
        return {}

def load_reservations_from_db():
    """Rezervasyonları veritabanından yükle"""
    try:
        conn = get_db_connection()
        if not conn:
            return []
            
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name_surname, center, venue, date, time, description, 
                   status, created_at, updated_at, created_by
            FROM reservations 
            ORDER BY created_at DESC
        """)
        db_reservations = cursor.fetchall()
        
        reservations_list = []
        for res in db_reservations:
            reservations_list.append({
                'id': res['id'],
                'name_surname': res['name_surname'],
                'center': res['center'],
                'venue': res['venue'] or 'Tiyatro Salonu',
                'date': res['date'].strftime('%Y-%m-%d'),
                'time': res['time'],
                'description': res['description'] or '',
                'status': res['status'],
                'created_at': res['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': res['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if res['updated_at'] else None,
                'created_by': res['created_by']
            })
        
        cursor.close()
        conn.close()
        return reservations_list
        
    except mysql.connector.Error as err:
        logger.error(f"Rezervasyon yükleme hatası: {err}")
        return []

def save_reservation_to_db(reservation_data):
    """Rezervasyonu veritabanına kaydet"""
    try:
        conn = get_db_connection()
        if not conn:
            return None
            
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO reservations (name_surname, center, venue, date, time, description, status, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            reservation_data['name_surname'],
            reservation_data['center'],
            reservation_data['venue'],
            reservation_data['date'],
            reservation_data['time'],
            reservation_data['description'],
            reservation_data['status'],
            reservation_data['created_by']
        ))
        
        reservation_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()
        
        logger.info(f"Rezervasyon kaydedildi: ID {reservation_id}")
        return reservation_id
        
    except mysql.connector.Error as err:
        logger.error(f"Rezervasyon kaydetme hatası: {err}")
        return None

def update_reservation_in_db(reservation_id, reservation_data):
    """Rezervasyonu veritabanında güncelle"""
    try:
        conn = get_db_connection()
        if not conn:
            return False
            
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE reservations 
            SET name_surname=%s, center=%s, venue=%s, date=%s, time=%s, description=%s
            WHERE id=%s
        """, (
            reservation_data['name_surname'],
            reservation_data['center'],
            reservation_data['venue'],
            reservation_data['date'],
            reservation_data['time'],
            reservation_data['description'],
            reservation_id
        ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        logger.info(f"Rezervasyon güncellendi: ID {reservation_id}")
        return True
        
    except mysql.connector.Error as err:
        logger.error(f"Rezervasyon güncelleme hatası: {err}")
        return False

def update_reservation_status_in_db(reservation_id, status):
    """Rezervasyon durumunu güncelle"""
    try:
        conn = get_db_connection()
        if not conn:
            return False
            
        cursor = conn.cursor()
        cursor.execute("UPDATE reservations SET status=%s WHERE id=%s", (status, reservation_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        logger.info(f"Rezervasyon durumu güncellendi: ID {reservation_id}, Durum: {status}")
        return True
        
    except mysql.connector.Error as err:
        logger.error(f"Rezervasyon durum güncelleme hatası: {err}")
        return False

def delete_reservation_from_db(reservation_id):
    """Rezervasyonu veritabanından sil"""
    try:
        conn = get_db_connection()
        if not conn:
            return False
            
        cursor = conn.cursor()
        cursor.execute("DELETE FROM reservations WHERE id=%s", (reservation_id,))
        
        affected_rows = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()
        
        if affected_rows > 0:
            logger.info(f"Rezervasyon silindi: ID {reservation_id}")
            return True
        return False
        
    except mysql.connector.Error as err:
        logger.error(f"Rezervasyon silme hatası: {err}")
        return False

def sync_reservations_to_memory():
    """Veritabanından rezervasyonları memory'ye yükle"""
    global reservations
    try:
        if connection_pool:
            reservations = load_reservations_from_db()
            logger.info(f"Rezervasyonlar senkronize edildi: {len(reservations)} kayıt")
            return True
        return False
    except Exception as e:
        logger.error(f"Rezervasyon senkronizasyon hatası: {e}")
        return False

# Uygulama başlatıldığında veritabanını ve kullanıcıları yükle
if connection_pool:
    init_database()
    users = load_users_from_db()
    sync_reservations_to_memory()  # Yeni senkronizasyon
    logger.info(f"Sistem başlatıldı. Kullanıcı sayısı: {len(users)}, Rezervasyon sayısı: {len(reservations)}")
else:
    # MySQL bağlantısı yoksa eski sistem
    users = {
        'admin': {
            'password': 'admin123', 
            'role': 'admin',
            'permissions': ['view_reservations', 'edit_reservations', 'view_availability', 'manage_users']
        },
        'editor1': {
            'password': 'editor123', 
            'role': 'editor',
            'permissions': ['edit_reservations']
        },
        'viewer1': {
            'password': 'viewer123', 
            'role': 'viewer',
            'permissions': ['view_reservations']
        },
        'scheduler1': {
            'password': 'scheduler123', 
            'role': 'scheduler',
            'permissions': ['view_availability']
        },
        'user1': {
            'password': 'user123', 
            'role': 'user',
            'permissions': []
        }
    }
    logger.warning("MySQL bağlantısı yok, memory-based sistem kullanılıyor")

def require_permission(permission):
    """Decorator: Belirli yetki gerektirir"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Bu sayfaya erişmek için giriş yapmalısınız.', 'error')
                return redirect(url_for('login'))
            
            user_id = session.get('user_id')
            user_permissions = users.get(user_id, {}).get('permissions', [])
            
            if permission not in user_permissions:
                flash('Bu işlem için gerekli yetkiniz bulunmamaktadır.', 'error')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def has_permission(permission):
    """Kullanıcının belirli yetkisi olup olmadığını kontrol et - Hiyerarşik sistem"""
    if 'user_id' not in session:
        return False
    
    user_id = session.get('user_id')
    user_permissions = users.get(user_id, {}).get('permissions', [])
    
    # Hiyerarşik yetki kontrolü
    if permission == 'view_reservations':
        # Rezervasyon listesi görüntüleme: edit_reservations yetkisi de bu yetkiyi verir
        return 'view_reservations' in user_permissions or 'edit_reservations' in user_permissions
    
    elif permission == 'view_availability':
        # Saat durumu görüntüleme: edit_reservations yetkisi de bu yetkiyi verir
        return 'view_availability' in user_permissions or 'edit_reservations' in user_permissions
    
    elif permission == 'edit_reservations':
        # Düzenleme yetkisi: sadece edit_reservations yetkisi olanlar
        return 'edit_reservations' in user_permissions
    
    elif permission == 'manage_users':
        # Kullanıcı yönetimi: sadece manage_users yetkisi olanlar
        return 'manage_users' in user_permissions
    
    else:
        return permission in user_permissions

def is_admin():
    """Mevcut kullanıcının admin olup olmadığını kontrol et"""
    return has_permission('manage_users')

def is_logged_in():
    """Kullanıcının giriş yapıp yapmadığını kontrol et"""
    return 'user_id' in session

def check_reservation_conflict(center, date, time, venue=None):
    """Rezervasyon çakışmasını kontrol et - Etkinlik yeri dahil"""
    for reservation in reservations:
        if (reservation['center'] == center and 
            reservation['date'] == date and 
            reservation['time'] == time and
            reservation.get('venue', 'Tiyatro Salonu') == venue and  # Varsayılan değer eski veriler için
            reservation['status'] in ['onay', 'bekle']):  # Sadece aktif rezervasyonlar
            return True
    return False

def get_alternative_times(center, date, selected_time, venue=None):
    """Alternatif saat dilimlerini öner - Etkinlik yeri dahil"""
    time_slots = [
        "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00",
        "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00",
        "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00", "21:00-22:00"
    ]
    
    try:
        current_index = time_slots.index(selected_time)
    except ValueError:
        return []
    
    alternatives = []
    
    # Bir saat öncesi
    if current_index > 0:
        prev_time = time_slots[current_index - 1]
        if not check_reservation_conflict(center, date, prev_time, venue):
            alternatives.append(prev_time)
    
    # Bir saat sonrası
    if current_index < len(time_slots) - 1:
        next_time = time_slots[current_index + 1]
        if not check_reservation_conflict(center, date, next_time, venue):
            alternatives.append(next_time)
    
    return alternatives

def get_filtered_reservations(center_filter=None, status_filter=None, month_filter=None, year_filter=None, venue_filter=None):
    """Rezervasyonları filtreleme - Etkinlik yeri dahil"""
    filtered = reservations.copy()
    
    if center_filter and center_filter != 'all':
        filtered = [r for r in filtered if r['center'] == center_filter]
    
    if status_filter and status_filter != 'all':
        filtered = [r for r in filtered if r['status'] == status_filter]
    
    if month_filter and month_filter != 'all':
        try:
            year, month = month_filter.split('-')
            filtered = [r for r in filtered if r['date'].startswith(f"{year}-{month}")]
        except ValueError:
            pass
    
    if year_filter and year_filter != 'all':
        filtered = [r for r in filtered if r['date'].startswith(f"{year_filter}-")]
    
    if venue_filter and venue_filter != 'all':
        filtered = [r for r in filtered if r.get('venue', 'Tiyatro Salonu') == venue_filter]
    
    return filtered

def get_available_months():
    """Mevcut rezervasyonların aylarını getir"""
    months = set()
    for reservation in reservations:
        date_parts = reservation['date'].split('-')
        if len(date_parts) >= 2:
            months.add(f"{date_parts[0]}-{date_parts[1]}")
    return sorted(list(months))

def get_available_years():
    """Mevcut rezervasyonların yıllarını getir + gelecek 5 yıl"""
    years = set()
    current_year = datetime.now().year
    
    # Mevcut rezervasyonlardan yılları al
    for reservation in reservations:
        date_parts = reservation['date'].split('-')
        if len(date_parts) >= 1:
            try:
                year = int(date_parts[0])
                years.add(year)
            except ValueError:
                pass
    
    # Mevcut yıl ve gelecek 5 yılı ekle
    for i in range(6):  # 2025, 2026, 2027, 2028, 2029, 2030
        years.add(current_year + i)
    
    return sorted(list(years))

def create_excel_file(reservations_data, filters):
    """Excel dosyası oluştur - Güvenli hata yönetimi ile"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rezervasyonlar"
        
        # Başlık stilleri
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Kenarlık stili
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık bilgileri
        ws.merge_cells('A1:I1')
        ws['A1'] = "REZERVASYON LİSTESİ"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Filtre bilgileri - Güvenli erişim
        filter_info = []
        if filters.get('center', 'all') != 'all':
            filter_info.append(f"Merkez: {filters.get('center', '')}")
        if filters.get('venue', 'all') != 'all':
            filter_info.append(f"Etkinlik Yeri: {filters.get('venue', '')}")
        if filters.get('status', 'all') != 'all':
            status_names = {'onay': 'Onaylı', 'bekle': 'Beklemede', 'iptal': 'İptal'}
            status_value = filters.get('status', '')
            filter_info.append(f"Durum: {status_names.get(status_value, status_value)}")
        if filters.get('month', 'all') != 'all':
            month_names = {
                '01': 'Ocak', '02': 'Şubat', '03': 'Mart', '04': 'Nisan',
                '05': 'Mayıs', '06': 'Haziran', '07': 'Temmuz', '08': 'Ağustos',
                '09': 'Eylül', '10': 'Ekim', '11': 'Kasım', '12': 'Aralık'
            }
            month_value = filters.get('month', '')
            if month_value and '-' in month_value:
                try:
                    year, month_num = month_value.split('-')
                    filter_info.append(f"Ay: {month_names.get(month_num, month_num)} {year}")
                except ValueError:
                    filter_info.append(f"Ay: {month_value}")
        if filters.get('year', 'all') != 'all':
            filter_info.append(f"Yıl: {filters.get('year', '')}")
        
        if filter_info:
            ws.merge_cells('A2:I2')
            ws['A2'] = f"Filtreler: {' | '.join(filter_info)}"
            ws['A2'].font = Font(italic=True)
            ws['A2'].alignment = Alignment(horizontal="center")
            start_row = 4
        else:
            start_row = 3
        
        # Tarih bilgisi
        ws.merge_cells(f'A{start_row-1}:I{start_row-1}')
        ws[f'A{start_row-1}'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws[f'A{start_row-1}'].font = Font(size=10)
        ws[f'A{start_row-1}'].alignment = Alignment(horizontal="right")
        
        # Tablo başlıkları
        headers = ['#', 'Ad Soyad', 'Merkez', 'Etkinlik Yeri', 'Tarih', 'Saat', 'Durum', 'Açıklama', 'Oluşturulma']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Veri satırları
        status_colors = {
            'onay': PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid"),
            'bekle': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            'iptal': PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        }
        
        status_names = {'onay': 'Onaylı', 'bekle': 'Beklemede', 'iptal': 'İptal'}
        
        for row_idx, reservation in enumerate(reservations_data, start_row + 1):
            # Güvenli veri erişimi
            try:
                # Tarih formatını düzenle
                date_value = reservation.get('date', '')
                try:
                    if isinstance(date_value, str):
                        date_obj = datetime.strptime(date_value, '%Y-%m-%d')
                        formatted_date = date_obj.strftime('%d.%m.%Y')
                    else:
                        formatted_date = str(date_value)
                except:
                    formatted_date = str(date_value)
                
                # Oluşturulma tarihini formatla
                created_value = reservation.get('created_at', '')
                try:
                    if isinstance(created_value, str) and created_value:
                        created_obj = datetime.strptime(created_value, '%Y-%m-%d %H:%M:%S')
                        formatted_created = created_obj.strftime('%d.%m.%Y %H:%M')
                    else:
                        formatted_created = str(created_value) if created_value else '-'
                except:
                    formatted_created = str(created_value) if created_value else '-'
                
                data = [
                    reservation.get('id', ''),
                    reservation.get('name_surname', ''),
                    reservation.get('center', ''),
                    reservation.get('venue', 'Tiyatro Salonu'),
                    formatted_date,
                    reservation.get('time', ''),
                    status_names.get(reservation.get('status', ''), reservation.get('status', '')),
                    reservation.get('description', '') or '-',
                    formatted_created
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = str(value) if value is not None else ''
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    
                    # Durum sütunu için renk
                    if col == 7:  # Durum sütunu
                        reservation_status = reservation.get('status', '')
                        if reservation_status in status_colors:
                            cell.fill = status_colors[reservation_status]
                            
            except Exception as e:
                logger.error(f"Excel satır hatası: {e}")
                continue
        
        # Sütun genişliklerini ayarla
        column_widths = [5, 20, 25, 15, 12, 15, 12, 30, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Satır yüksekliklerini ayarla
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        return wb
        
    except Exception as e:
        logger.error(f"Excel dosyası oluşturma hatası: {e}")
        # Hata durumunda basit bir Excel dosyası oluştur
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Hata: Excel dosyası oluşturulamadı"
        return wb

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Kullanıcı giriş sayfası"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if username in users and users[username]['password'] == password:
            session['user_id'] = username
            session['user_role'] = users[username]['role']
            session['user_permissions'] = users[username]['permissions']
            
            flash(f'Hoş geldiniz, {username}!', 'success')
            
            # Yetkilere göre yönlendirme
            if has_permission('view_reservations'):
                return redirect(url_for('reservations_list'))
            elif has_permission('view_availability'):
                return redirect(url_for('availability_view'))
            else:
                return redirect(url_for('index'))
        else:
            flash('Kullanıcı adı veya şifre hatalı!', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Kullanıcı çıkış"""
    session.clear()
    flash('Başarıyla çıkış yaptınız.', 'success')
    return redirect(url_for('login'))

@app.route('/', methods=['GET', 'POST'])
def index():
    if not is_logged_in():
        return redirect(url_for('login'))
        
    if request.method == 'POST':
        # Form verilerini al ve doğrula
        name_surname = request.form.get('name_surname', '').strip()
        center = request.form.get('center', '')
        venue = request.form.get('venue', '')
        date = request.form.get('date', '')
        time = request.form.get('time', '')
        description = request.form.get('description', '').strip()

        # Temel validasyon
        if not all([name_surname, center, venue, date, time]):
            flash('Lütfen tüm zorunlu alanları doldurunuz!', 'error')
            return render_template('index.html', form_data={
                'name_surname': name_surname,
                'center': center,
                'venue': venue,
                'date': date,
                'time': time,
                'description': description
            })

        # Çakışma kontrolü - Etkinlik yeri dahil
        if check_reservation_conflict(center, date, time, venue):
            alternatives = get_alternative_times(center, date, time, venue)
            
            if alternatives:
                alt_text = ", ".join(alternatives)
                flash(f'Bu saat dilimi ({time}) için {center} - {venue}\'nde {date} tarihinde zaten rezervasyon var! Alternatif saatler: {alt_text}', 'error')
            else:
                flash(f'Bu saat dilimi ({time}) için {center} - {venue}\'nde {date} tarihinde zaten rezervasyon var ve alternatif saat bulunamadı!', 'error')
            
            return render_template('index.html', form_data={
                'name_surname': name_surname,
                'center': center,
                'venue': venue,
                'date': date,
                'time': time,
                'description': description
            })

        # Rezervasyon verisini hazırla
        reservation_data = {
            'name_surname': name_surname,
            'center': center,
            'venue': venue,
            'date': date,
            'time': time,
            'description': description,
            'status': 'bekle',
            'created_by': session.get('user_id')
        }

        # MySQL'e kaydet
        if connection_pool:
            reservation_id = save_reservation_to_db(reservation_data)
            if reservation_id:
                # Memory'ye de ekle (anlık sync için)
                reservation_data.update({
                    'id': reservation_id,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                reservations.append(reservation_data)
                flash('Rezervasyon başarıyla oluşturuldu! Durum: Beklemede', 'success')
            else:
                flash('Rezervasyon kaydedilirken hata oluştu!', 'error')
                return render_template('index.html', form_data=reservation_data)
        else:
            # MySQL yoksa eski sistem
            reservation_data.update({
                'id': len(reservations) + 1,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            reservations.append(reservation_data)
            flash('Rezervasyon başarıyla oluşturuldu! Durum: Beklemede', 'success')
        
        # Yönlendirme
        if is_admin():
            return redirect(url_for('reservations_list'))
        else:
            return render_template('index.html', success_message=True)

    return render_template('index.html')

@app.route('/reservations')
def reservations_list():
    """Rezervasyon listesi - Hiyerarşik yetki kontrolü"""
    if not has_permission('view_reservations'):
        flash('Bu sayfaya erişmek için gerekli yetkiniz bulunmamaktadır.', 'error')
        return redirect(url_for('index'))
    
    # Filtre parametrelerini al
    center_filter = request.args.get('center', 'all')
    status_filter = request.args.get('status', 'all')
    month_filter = request.args.get('month', 'all')
    year_filter = request.args.get('year', str(datetime.now().year))
    venue_filter = request.args.get('venue', 'all')  # Yeni filtre
    
    # Filtrelenmiş rezervasyonları getir
    filtered_reservations = get_filtered_reservations(center_filter, status_filter, month_filter, year_filter, venue_filter)
    
    # Filtre seçenekleri için veriler
    centers = ['Sefaköy Kültür Merkezi', 'Cennet Kültür Merkezi', 'Atakent Kültür Merkezi', 'Kemalpaşa Semt Konağı']
    venues = ['Tiyatro Salonu', 'Seminer Salonu']  # Yeni seçenekler
    statuses = [
        {'value': 'onay', 'label': 'Onaylı'},
        {'value': 'bekle', 'label': 'Beklemede'},
        {'value': 'iptal', 'label': 'İptal'}
    ]
    available_months = get_available_months()
    available_years = get_available_years()
    
    return render_template('reservations.html', 
                         reservations=filtered_reservations,
                         centers=centers,
                         venues=venues,
                         statuses=statuses,
                         available_months=available_months,
                         available_years=available_years,
                         current_filters={
                             'center': center_filter,
                             'status': status_filter,
                             'month': month_filter,
                             'year': year_filter,
                             'venue': venue_filter
                         })

@app.route('/export/excel')
def export_excel():
    """Excel dosyası olarak export et - Yetki kontrolü güncellenecek"""
    # view_reservations yetkisi olanlar Excel indirebilir
    if not has_permission('view_reservations'):
        flash('Bu işlem için gerekli yetkiniz bulunmamaktadır.', 'error')
        return redirect(url_for('index'))
    
    # Filtre parametrelerini güvenli şekilde al
    center_filter = request.args.get('center', 'all')
    status_filter = request.args.get('status', 'all')
    month_filter = request.args.get('month', 'all')
    year_filter = request.args.get('year', 'all')
    venue_filter = request.args.get('venue', 'all')  # Yeni filtre
    
    # Filtrelenmiş rezervasyonları getir
    try:
        filtered_reservations = get_filtered_reservations(center_filter, status_filter, month_filter, year_filter, venue_filter)
    except Exception as e:
        logger.error(f"Filtreleme hatası: {e}")
        filtered_reservations = reservations  # Hata durumunda tüm rezervasyonları al
    
    if not filtered_reservations:
        flash('Export edilecek rezervasyon bulunamadı!', 'warning')
        return redirect(url_for('reservations_list'))
    
    # Excel dosyası oluştur - Güvenli filtre objesi
    filters = {
        'center': center_filter if center_filter else 'all',
        'status': status_filter if status_filter else 'all',
        'month': month_filter if month_filter else 'all',
        'year': year_filter if year_filter else 'all',
        'venue': venue_filter if venue_filter else 'all'
    }
    
    try:
        wb = create_excel_file(filtered_reservations, filters)
        
        # Dosyayı memory'de oluştur
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Dosya adını oluştur
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"rezervasyonlar_{timestamp}.xlsx"
        
        # Response oluştur
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        flash(f'{len(filtered_reservations)} rezervasyon Excel dosyası olarak indirildi!', 'success')
        
        return response
        
    except Exception as e:
        logger.error(f"Excel oluşturma hatası: {e}")
        flash('Excel dosyası oluşturulurken hata oluştu!', 'error')
        return redirect(url_for('reservations_list'))

@app.route('/reservation/approve/<int:reservation_id>')
@require_permission('edit_reservations')
def approve_reservation(reservation_id):
    """Rezervasyonu onayla"""
    # MySQL'de güncelle
    if connection_pool and update_reservation_status_in_db(reservation_id, 'onay'):
        # Memory'de de güncelle
        for reservation in reservations:
            if reservation['id'] == reservation_id:
                reservation['status'] = 'onay'
                break
        flash(f"#{reservation_id} numaralı rezervasyon onaylandı!", 'success')
    else:
        # MySQL yoksa sadece memory
        for reservation in reservations:
            if reservation['id'] == reservation_id:
                reservation['status'] = 'onay'
                flash(f"#{reservation_id} numaralı rezervasyon onaylandı!", 'success')
                break
        else:
            flash('Rezervasyon bulunamadı!', 'error')
    return redirect(url_for('reservations_list'))

@app.route('/reservation/pending/<int:reservation_id>')
@require_permission('edit_reservations')
def pending_reservation(reservation_id):
    """Rezervasyonu beklemeye al"""
    # MySQL'de güncelle
    if connection_pool and update_reservation_status_in_db(reservation_id, 'bekle'):
        # Memory'de de güncelle
        for reservation in reservations:
            if reservation['id'] == reservation_id:
                reservation['status'] = 'bekle'
                break
        flash(f"#{reservation_id} numaralı rezervasyon beklemeye alındı!", 'warning')
    else:
        # MySQL yoksa sadece memory
        for reservation in reservations:
            if reservation['id'] == reservation_id:
                reservation['status'] = 'bekle'
                flash(f"#{reservation_id} numaralı rezervasyon beklemeye alındı!", 'warning')
                break
        else:
            flash('Rezervasyon bulunamadı!', 'error')
    return redirect(url_for('reservations_list'))

@app.route('/reservation/cancel/<int:reservation_id>')
@require_permission('edit_reservations')
def cancel_reservation(reservation_id):
    """Rezervasyonu iptal et"""
    # MySQL'de güncelle
    if connection_pool and update_reservation_status_in_db(reservation_id, 'iptal'):
        # Memory'de de güncelle
        for reservation in reservations:
            if reservation['id'] == reservation_id:
                reservation['status'] = 'iptal'
                break
        flash(f"#{reservation_id} numaralı rezervasyon iptal edildi!", 'warning')
    else:
        # MySQL yoksa sadece memory
        for reservation in reservations:
            if reservation['id'] == reservation_id:
                reservation['status'] = 'iptal'
                flash(f"#{reservation_id} numaralı rezervasyon iptal edildi!", 'warning')
                break
        else:
            flash('Rezervasyon bulunamadı!', 'error')
    return redirect(url_for('reservations_list'))

@app.route('/reservation/delete/<int:reservation_id>')
@require_permission('edit_reservations')
def delete_reservation(reservation_id):
    """Rezervasyonu sil"""
    global reservations
    
    # MySQL'den sil
    if connection_pool and delete_reservation_from_db(reservation_id):
        # Memory'den de sil
        initial_length = len(reservations)
        reservations = [r for r in reservations if r['id'] != reservation_id]
        flash(f"#{reservation_id} numaralı rezervasyon silindi!", 'success')
    else:
        # MySQL yoksa sadece memory
        initial_length = len(reservations)
        reservations = [r for r in reservations if r['id'] != reservation_id]
        
        if len(reservations) < initial_length:
            flash(f"#{reservation_id} numaralı rezervasyon silindi!", 'success')
        else:
            flash('Rezervasyon bulunamadı!', 'error')
    
    return redirect(url_for('reservations_list'))

@app.route('/reservation/edit/<int:reservation_id>')
@require_permission('edit_reservations')
def edit_reservation(reservation_id):
    """Rezervasyonu düzenleme sayfası"""
    for reservation in reservations:
        if reservation['id'] == reservation_id:
            return render_template('edit_reservation.html', reservation=reservation)
    
    flash('Rezervasyon bulunamadı!', 'error')
    return redirect(url_for('reservations_list'))

@app.route('/reservation/update/<int:reservation_id>', methods=['POST'])
@require_permission('edit_reservations')
def update_reservation(reservation_id):
    """Rezervasyonu güncelle"""
    # Form verilerini al
    name_surname = request.form.get('name_surname', '').strip()
    center = request.form.get('center', '')
    venue = request.form.get('venue', '')
    date = request.form.get('date', '')
    time = request.form.get('time', '')
    description = request.form.get('description', '').strip()

    # Rezervasyonu bul
    target_reservation = None
    for reservation in reservations:
        if reservation['id'] == reservation_id:
            target_reservation = reservation
            break
    
    if not target_reservation:
        flash('Rezervasyon bulunamadı!', 'error')
        return redirect(url_for('reservations_list'))

    # Temel validasyon
    if not all([name_surname, center, venue, date, time]):
        flash('Lütfen tüm zorunlu alanları doldurunuz!', 'error')
        return render_template('edit_reservation.html', 
                             reservation=target_reservation,
                             form_data={
                                 'name_surname': name_surname,
                                 'center': center,
                                 'venue': venue,
                                 'date': date,
                                 'time': time,
                                 'description': description
                             })

    # Çakışma kontrolü (kendisi hariç)
    conflict_found = False
    for reservation in reservations:
        if (reservation['id'] != reservation_id and
            reservation['center'] == center and 
            reservation['date'] == date and 
            reservation['time'] == time and
            reservation.get('venue', 'Tiyatro Salonu') == venue and
            reservation['status'] in ['onay', 'bekle']):
            conflict_found = True
            break

    if conflict_found:
        alternatives = get_alternative_times(center, date, time, venue)
        if alternatives:
            alt_text = ", ".join(alternatives)
            flash(f'Bu saat dilimi ({time}) için {center} - {venue}\'nde {date} tarihinde zaten rezervasyon var! Alternatif saatler: {alt_text}', 'error')
        else:
            flash(f'Bu saat dilimi ({time}) için {center} - {venue}\'nde {date} tarihinde zaten rezervasyon var ve alternatif saat bulunamadı!', 'error')
        
        return render_template('edit_reservation.html', 
                             reservation=target_reservation,
                             form_data={
                                 'name_surname': name_surname,
                                 'center': center,
                                 'venue': venue,
                                 'date': date,
                                 'time': time,
                                 'description': description
                             })

    # Güncelleme verilerini hazırla
    update_data = {
        'name_surname': name_surname,
        'center': center,
        'venue': venue,
        'date': date,
        'time': time,
        'description': description
    }

    # MySQL'de güncelle
    if connection_pool and update_reservation_in_db(reservation_id, update_data):
        # Memory'de de güncelle
        target_reservation.update(update_data)
        target_reservation['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        flash(f'#{reservation_id} numaralı rezervasyon başarıyla güncellendi!', 'success')
    else:
        # MySQL yoksa sadece memory
        target_reservation.update(update_data)
        target_reservation['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        flash(f'#{reservation_id} numaralı rezervasyon başarıyla güncellendi!', 'success')

    return redirect(url_for('reservations_list'))

def get_time_availability(center, date, venue=None):
    """Belirli bir merkez, tarih ve etkinlik yeri için saat durumlarını getir"""
    time_slots = [
        "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00",
        "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00",
        "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00", "21:00-22:00"
    ]
    
    availability = {}
    for slot in time_slots:
        is_occupied = False
        for reservation in reservations:
            if (reservation['center'] == center and 
                reservation['date'] == date and 
                reservation['time'] == slot and
                reservation.get('venue', 'Tiyatro Salonu') == venue and
                reservation['status'] in ['onay', 'bekle']):
                is_occupied = True
                break
        availability[slot] = 'dolu' if is_occupied else 'bos'
    
    return availability

@app.route('/availability')
def availability_view():
    """Saat durumu görüntüleme sayfası - Hiyerarşik yetki kontrolü"""
    if not has_permission('view_availability'):
        flash('Bu sayfaya erişmek için gerekli yetkiniz bulunmamaktadır.', 'error')
        return redirect(url_for('index'))
    
    selected_center = request.args.get('center', '')
    selected_date = request.args.get('date', '')
    selected_venue = request.args.get('venue', 'Tiyatro Salonu')  # Yeni parametre
    
    centers = ['Sefaköy Kültür Merkezi', 'Cennet Kültür Merkezi', 'Atakent Kültür Merkezi', 'Kemalpaşa Semt Konağı']
    venues = ['Tiyatro Salonu', 'Seminer Salonu']  # Yeni seçenekler
    availability_data = {}
    
    if selected_center and selected_date and selected_venue:
        availability_data = get_time_availability(selected_center, selected_date, selected_venue)
    
    return render_template('availability.html', 
                         centers=centers,
                         venues=venues,
                         selected_center=selected_center,
                         selected_date=selected_date,
                         selected_venue=selected_venue,
                         availability_data=availability_data)

@app.route('/admin/users')
@require_permission('manage_users')
def admin_users():
    """Kullanıcı yönetimi sayfası"""
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/<username>/permissions', methods=['POST'])
@require_permission('manage_users')
def update_user_permissions(username):
    """Kullanıcı yetkilerini güncelle"""
    if username not in users:
        flash('Kullanıcı bulunamadı!', 'error')
        return redirect(url_for('admin_users'))
    
    # Yeni yetkileri al
    new_permissions = request.form.getlist('permissions')
    
    # Geçerli yetkileri kontrol et
    valid_permissions = ['view_reservations', 'edit_reservations', 'view_availability', 'manage_users']
    filtered_permissions = [p for p in new_permissions if p in valid_permissions]
    
    # Yetkileri güncelle
    users[username]['permissions'] = filtered_permissions
    
    flash(f'{username} kullanıcısının yetkileri güncellendi!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/add', methods=['POST'])
@require_permission('manage_users')
def add_user():
    """Yeni kullanıcı ekle"""
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role', 'user')
    
    if not username or not password:
        flash('Kullanıcı adı ve şifre gereklidir!', 'error')
        return redirect(url_for('admin_users'))
    
    if username in users:
        flash('Bu kullanıcı adı zaten kullanılıyor!', 'error')
        return redirect(url_for('admin_users'))
    
    # Yeni kullanıcı oluştur
    users[username] = {
        'password': password,
        'role': role,
        'permissions': []
    }
    
    flash(f'{username} kullanıcısı başarıyla eklendi!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<username>/delete')
@require_permission('manage_users')
def delete_user(username):
    """Kullanıcı sil"""
    if username == session.get('user_id'):
        flash('Kendi hesabınızı silemezsiniz!', 'error')
        return redirect(url_for('admin_users'))
    
    if username in users:
        del users[username]
        flash(f'{username} kullanıcısı silindi!', 'success')
    else:
        flash('Kullanıcı bulunamadı!', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<username>/reset-password')
@require_permission('manage_users')
def reset_user_password(username):
    """Kullanıcı şifresini varsayılan şifreye sıfırla"""
    if username == session.get('user_id'):
        flash('Kendi şifrenizi bu şekilde sıfırlayamazsınız!', 'error')
        return redirect(url_for('admin_users'))
    
    if username not in users:
        flash('Kullanıcı bulunamadı!', 'error')
        return redirect(url_for('admin_users'))
    
    # Şifreyi varsayılan şifreye sıfırla
    default_password = '123456'
    users[username]['password'] = default_password
    
    flash(f'{username} kullanıcısının şifresi varsayılan şifreye sıfırlandı! (Yeni şifre: {default_password})', 'success')
    return redirect(url_for('admin_users'))

if __name__ == '__main__':
    # Environment variables'dan Flask ayarları
    debug_mode = os.getenv('FLASK_DEBUG', 'True').lower() == 'true'
    host = os.getenv('FLASK_HOST', '0.0.0.0')
    port = int(os.getenv('FLASK_PORT', 5001))
    
    app.run(debug=debug_mode, host=host, port=port)  # Port değiştir