from flask import Flask, render_template, request, redirect, url_for, flash, make_response
import secrets
from datetime import datetime
import io
import mysql.connector
from mysql.connector import Error
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)

# MySQL Veritabanı Yapılandırması
DB_CONFIG = {
    'host': 'localhost',
    'database': 'rezervasyon_sistemi',
    'user': 'root',
    'password': 'password',  # Kendi şifrenizi buraya yazın
    'port': 3306,
    'charset': 'utf8mb4'
}

def get_db_connection():
    """MySQL veritabanı bağlantısı oluştur"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        print(f"MySQL bağlantı hatası: {e}")
        return None

def init_database():
    """Veritabanı ve tabloları oluştur"""
    try:
        connection = mysql.connector.connect(
            host=DB_CONFIG['host'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            port=DB_CONFIG['port']
        )
        cursor = connection.cursor()
        
        # Veritabanını oluştur
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_CONFIG['database']} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        cursor.execute(f"USE {DB_CONFIG['database']}")
        
        # Rezervasyonlar tablosunu oluştur
        create_table_query = """
        CREATE TABLE IF NOT EXISTS reservations (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name_surname VARCHAR(255) NOT NULL,
            center VARCHAR(255) NOT NULL,
            date DATE NOT NULL,
            time VARCHAR(20) NOT NULL,
            description TEXT,
            status ENUM('onay', 'bekle', 'iptal') DEFAULT 'bekle',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_center_date_time (center, date, time),
            INDEX idx_status (status),
            INDEX idx_date (date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
        cursor.execute(create_table_query)
        
        connection.commit()
        print("Veritabanı ve tablolar başarıyla oluşturuldu!")
        
    except Error as e:
        print(f"Veritabanı oluşturma hatası: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def check_reservation_conflict(center, date, time, exclude_id=None):
    """Rezervasyon çakışmasını kontrol et"""
    connection = get_db_connection()
    if not connection:
        return False
    
    try:
        cursor = connection.cursor()
        if exclude_id:
            query = """
            SELECT COUNT(*) FROM reservations 
            WHERE center = %s AND date = %s AND time = %s 
            AND status IN ('onay', 'bekle') AND id != %s
            """
            cursor.execute(query, (center, date, time, exclude_id))
        else:
            query = """
            SELECT COUNT(*) FROM reservations 
            WHERE center = %s AND date = %s AND time = %s 
            AND status IN ('onay', 'bekle')
            """
            cursor.execute(query, (center, date, time))
        
        result = cursor.fetchone()
        return result[0] > 0
        
    except Error as e:
        print(f"Çakışma kontrolü hatası: {e}")
        return False
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def get_alternative_times(center, date, selected_time):
    """Alternatif saat dilimlerini öner"""
    time_slots = [
        "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00",
        "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00",
        "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00", "21:00-22:00"
    ]
    
    try:
        current_index = time_slots.index(selected_time)
    except ValueError:
        return []
    
    alternatives = []
    
    # Bir saat öncesi
    if current_index > 0:
        prev_time = time_slots[current_index - 1]
        if not check_reservation_conflict(center, date, prev_time):
            alternatives.append(prev_time)
    
    # Bir saat sonrası
    if current_index < len(time_slots) - 1:
        next_time = time_slots[current_index + 1]
        if not check_reservation_conflict(center, date, next_time):
            alternatives.append(next_time)
    
    return alternatives

def get_filtered_reservations(center_filter=None, status_filter=None, month_filter=None, year_filter=None):
    """Rezervasyonları filtreleme"""
    connection = get_db_connection()
    if not connection:
        return []
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        query = "SELECT * FROM reservations WHERE 1=1"
        params = []
        
        if center_filter and center_filter != 'all':
            query += " AND center = %s"
            params.append(center_filter)
        
        if status_filter and status_filter != 'all':
            query += " AND status = %s"
            params.append(status_filter)
        
        if month_filter and month_filter != 'all':
            try:
                year, month = month_filter.split('-')
                query += " AND YEAR(date) = %s AND MONTH(date) = %s"
                params.extend([year, month])
            except ValueError:
                pass
        
        if year_filter and year_filter != 'all':
            query += " AND YEAR(date) = %s"
            params.append(year_filter)
        
        query += " ORDER BY date DESC, time ASC"
        
        cursor.execute(query, params)
        reservations = cursor.fetchall()
        
        # Tarih formatını string'e çevir
        for reservation in reservations:
            if reservation['date']:
                reservation['date'] = reservation['date'].strftime('%Y-%m-%d')
            if reservation['created_at']:
                reservation['created_at'] = reservation['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if reservation['updated_at']:
                reservation['updated_at'] = reservation['updated_at'].strftime('%Y-%m-%d %H:%M:%S')
        
        return reservations
        
    except Error as e:
        print(f"Rezervasyon listeleme hatası: {e}")
        return []
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def get_available_months():
    """Mevcut rezervasyonların aylarını getir"""
    connection = get_db_connection()
    if not connection:
        return []
    
    try:
        cursor = connection.cursor()
        query = """
        SELECT DISTINCT DATE_FORMAT(date, '%Y-%m') as month_year 
        FROM reservations 
        ORDER BY month_year DESC
        """
        cursor.execute(query)
        months = [row[0] for row in cursor.fetchall()]
        return months
        
    except Error as e:
        print(f"Ay listeleme hatası: {e}")
        return []
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def get_available_years():
    """Mevcut rezervasyonların yıllarını getir + gelecek 5 yıl"""
    connection = get_db_connection()
    years = set()
    current_year = datetime.now().year
    
    if connection:
        try:
            cursor = connection.cursor()
            query = "SELECT DISTINCT YEAR(date) as year FROM reservations"
            cursor.execute(query)
            db_years = [row[0] for row in cursor.fetchall()]
            years.update(db_years)
        except Error as e:
            print(f"Yıl listeleme hatası: {e}")
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    
    # Mevcut yıl ve gelecek 5 yılı ekle
    for i in range(6):
        years.add(current_year + i)
    
    return sorted(list(years))

def create_excel_file(reservations_data, filters):
    """Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rezervasyonlar"
    
    # Başlık stilleri
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Kenarlık stili
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlık bilgileri
    ws.merge_cells('A1:H1')
    ws['A1'] = "REZERVASYON LİSTESİ"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Filtre bilgileri
    filter_info = []
    if filters.get('center') != 'all':
        filter_info.append(f"Merkez: {filters['center']}")
    if filters.get('status') != 'all':
        status_names = {'onay': 'Onaylı', 'bekle': 'Beklemede', 'iptal': 'İptal'}
        filter_info.append(f"Durum: {status_names.get(filters['status'], filters['status'])}")
    if filters.get('month') != 'all':
        month_names = {
            '01': 'Ocak', '02': 'Şubat', '03': 'Mart', '04': 'Nisan',
            '05': 'Mayıs', '06': 'Haziran', '07': 'Temmuz', '08': 'Ağustos',
            '09': 'Eylül', '10': 'Ekim', '11': 'Kasım', '12': 'Aralık'
        }
        year, month_num = filters['month'].split('-')
        filter_info.append(f"Ay: {month_names[month_num]} {year}")
    if filters.get('year') != 'all':
        filter_info.append(f"Yıl: {filters['year']}")
    
    if filter_info:
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Filtreler: {' | '.join(filter_info)}"
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = Alignment(horizontal="center")
        start_row = 4
    else:
        start_row = 3
    
    # Tarih bilgisi
    ws.merge_cells(f'A{start_row-1}:H{start_row-1}')
    ws[f'A{start_row-1}'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws[f'A{start_row-1}'].font = Font(size=10)
    ws[f'A{start_row-1}'].alignment = Alignment(horizontal="right")
    
    # Tablo başlıkları
    headers = ['#', 'Ad Soyad', 'Merkez', 'Tarih', 'Saat', 'Durum', 'Açıklama', 'Oluşturulma']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Veri satırları
    status_colors = {
        'onay': PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid"),
        'bekle': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        'iptal': PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    }
    
    status_names = {'onay': 'Onaylı', 'bekle': 'Beklemede', 'iptal': 'İptal'}
    
    for row_idx, reservation in enumerate(reservations_data, start_row + 1):
        # Tarih formatını düzenle
        try:
            date_obj = datetime.strptime(reservation['date'], '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d.%m.%Y')
        except:
            formatted_date = reservation['date']
        
        # Oluşturulma tarihini formatla
        try:
            created_obj = datetime.strptime(reservation['created_at'], '%Y-%m-%d %H:%M:%S')
            formatted_created = created_obj.strftime('%d.%m.%Y %H:%M')
        except:
            formatted_created = reservation.get('created_at', '-')
        
        data = [
            reservation['id'],
            reservation['name_surname'],
            reservation['center'],
            formatted_date,
            reservation['time'],
            status_names.get(reservation['status'], reservation['status']),
            reservation.get('description', '') or '-',
            formatted_created
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            # Durum sütunu için renk
            if col == 6:  # Durum sütunu
                if reservation['status'] in status_colors:
                    cell.fill = status_colors[reservation['status']]
    
    # Sütun genişliklerini ayarla
    column_widths = [5, 20, 25, 12, 15, 12, 30, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Satır yüksekliklerini ayarla
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
    
    return wb

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Form verilerini al ve doğrula
        name_surname = request.form.get('name_surname', '').strip()
        center = request.form.get('center', '')
        date = request.form.get('date', '')
        time = request.form.get('time', '')
        description = request.form.get('description', '').strip()

        # Temel validasyon
        if not all([name_surname, center, date, time]):
            flash('Lütfen tüm zorunlu alanları doldurunuz!', 'error')
            return render_template('index.html', form_data={
                'name_surname': name_surname,
                'center': center,
                'date': date,
                'time': time,
                'description': description
            })

        # Çakışma kontrolü
        if check_reservation_conflict(center, date, time):
            alternatives = get_alternative_times(center, date, time)
            
            if alternatives:
                alt_text = ", ".join(alternatives)
                flash(f'Bu saat dilimi ({time}) için {center} merkezinde {date} tarihinde zaten rezervasyon var! Alternatif saatler: {alt_text}', 'error')
            else:
                flash(f'Bu saat dilimi ({time}) için {center} merkezinde {date} tarihinde zaten rezervasyon var ve alternatif saat bulunamadı!', 'error')
            
            return render_template('index.html', form_data={
                'name_surname': name_surname,
                'center': center,
                'date': date,
                'time': time,
                'description': description
            })

        # Veritabanına rezervasyon ekle
        connection = get_db_connection()
        if connection:
            try:
                cursor = connection.cursor()
                query = """
                INSERT INTO reservations (name_surname, center, date, time, description, status)
                VALUES (%s, %s, %s, %s, %s, 'bekle')
                """
                cursor.execute(query, (name_surname, center, date, time, description))
                connection.commit()
                
                flash('Rezervasyon başarıyla oluşturuldu! Durum: Beklemede', 'success')
                return redirect(url_for('reservations_list'))
                
            except Error as e:
                flash(f'Rezervasyon oluşturulurken hata oluştu: {e}', 'error')
                return render_template('index.html', form_data={
                    'name_surname': name_surname,
                    'center': center,
                    'date': date,
                    'time': time,
                    'description': description
                })
            finally:
                if connection.is_connected():
                    cursor.close()
                    connection.close()
        else:
            flash('Veritabanı bağlantısı kurulamadı!', 'error')

    return render_template('index.html')

@app.route('/reservations')
def reservations_list():
    # Filtre parametrelerini al
    center_filter = request.args.get('center', 'all')
    status_filter = request.args.get('status', 'all')
    month_filter = request.args.get('month', 'all')
    year_filter = request.args.get('year', str(datetime.now().year))
    
    # Filtrelenmiş rezervasyonları getir
    filtered_reservations = get_filtered_reservations(center_filter, status_filter, month_filter, year_filter)
    
    # Filtre seçenekleri için veriler
    centers = ['Sefaköy Kültür Merkezi', 'Cennet Kültür Merkezi', 'Atakent Kültür Merkezi', 'Kemalpaşa Semt Konağı']
    statuses = [
        {'value': 'onay', 'label': 'Onaylı'},
        {'value': 'bekle', 'label': 'Beklemede'},
        {'value': 'iptal', 'label': 'İptal'}
    ]
    available_months = get_available_months()
    available_years = get_available_years()
    
    return render_template('reservations.html', 
                         reservations=filtered_reservations,
                         centers=centers,
                         statuses=statuses,
                         available_months=available_months,
                         available_years=available_years,
                         current_filters={
                             'center': center_filter,
                             'status': status_filter,
                             'month': month_filter,
                             'year': year_filter
                         })

@app.route('/reservation/approve/<int:reservation_id>')
def approve_reservation(reservation_id):
    """Rezervasyonu onayla"""
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            query = "UPDATE reservations SET status = 'onay' WHERE id = %s"
            cursor.execute(query, (reservation_id,))
            connection.commit()
            
            if cursor.rowcount > 0:
                flash(f"#{reservation_id} numaralı rezervasyon onaylandı!", 'success')
            else:
                flash('Rezervasyon bulunamadı!', 'error')
                
        except Error as e:
            flash(f'Rezervasyon onaylanırken hata oluştu: {e}', 'error')
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    else:
        flash('Veritabanı bağlantısı kurulamadı!', 'error')
    
    return redirect(url_for('reservations_list'))

@app.route('/reservation/pending/<int:reservation_id>')
def pending_reservation(reservation_id):
    """Rezervasyonu beklemeye al"""
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            query = "UPDATE reservations SET status = 'bekle' WHERE id = %s"
            cursor.execute(query, (reservation_id,))
            connection.commit()
            
            if cursor.rowcount > 0:
                flash(f"#{reservation_id} numaralı rezervasyon beklemeye alındı!", 'warning')
            else:
                flash('Rezervasyon bulunamadı!', 'error')
                
        except Error as e:
            flash(f'Rezervasyon güncellenirken hata oluştu: {e}', 'error')
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    else:
        flash('Veritabanı bağlantısı kurulamadı!', 'error')
    
    return redirect(url_for('reservations_list'))

@app.route('/reservation/cancel/<int:reservation_id>')
def cancel_reservation(reservation_id):
    """Rezervasyonu iptal et"""
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            query = "UPDATE reservations SET status = 'iptal' WHERE id = %s"
            cursor.execute(query, (reservation_id,))
            connection.commit()
            
            if cursor.rowcount > 0:
                flash(f"#{reservation_id} numaralı rezervasyon iptal edildi!", 'warning')
            else:
                flash('Rezervasyon bulunamadı!', 'error')
                
        except Error as e:
            flash(f'Rezervasyon iptal edilirken hata oluştu: {e}', 'error')
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    else:
        flash('Veritabanı bağlantısı kurulamadı!', 'error')
    
    return redirect(url_for('reservations_list'))

@app.route('/reservation/delete/<int:reservation_id>')
def delete_reservation(reservation_id):
    """Rezervasyonu sil"""
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            query = "DELETE FROM reservations WHERE id = %s"
            cursor.execute(query, (reservation_id,))
            connection.commit()
            
            if cursor.rowcount > 0:
                flash(f"#{reservation_id} numaralı rezervasyon silindi!", 'success')
            else:
                flash('Rezervasyon bulunamadı!', 'error')
                
        except Error as e:
            flash(f'Rezervasyon silinirken hata oluştu: {e}', 'error')
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    else:
        flash('Veritabanı bağlantısı kurulamadı!', 'error')
    
    return redirect(url_for('reservations_list'))

@app.route('/export/excel')
def export_excel():
    """Excel dosyası olarak export et"""
    # Filtre parametrelerini al
    center_filter = request.args.get('center', 'all')
    status_filter = request.args.get('status', 'all')
    month_filter = request.args.get('month', 'all')
    year_filter = request.args.get('year', 'all')
    
    # Filtrelenmiş rezervasyonları getir
    filtered_reservations = get_filtered_reservations(center_filter, status_filter, month_filter, year_filter)
    
    if not filtered_reservations:
        flash('Export edilecek rezervasyon bulunamadı!', 'warning')
        return redirect(url_for('reservations_list'))
    
    # Excel dosyası oluştur
    filters = {
        'center': center_filter,
        'status': status_filter,
        'month': month_filter,
        'year': year_filter
    }
    
    wb = create_excel_file(filtered_reservations, filters)
    
    # Dosyayı memory'de oluştur
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Dosya adını oluştur
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"rezervasyonlar_{timestamp}.xlsx"
    
    # Response oluştur
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    flash(f'{len(filtered_reservations)} rezervasyon Excel dosyası olarak indirildi!', 'success')
    
    return response

@app.route('/reservation/edit/<int:reservation_id>')
def edit_reservation(reservation_id):
    """Rezervasyonu düzenleme sayfası"""
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor(dictionary=True)
            query = "SELECT * FROM reservations WHERE id = %s"
            cursor.execute(query, (reservation_id,))
            reservation = cursor.fetchone()
            
            if reservation:
                # Tarih formatını string'e çevir
                if reservation['date']:
                    reservation['date'] = reservation['date'].strftime('%Y-%m-%d')
                if reservation['created_at']:
                    reservation['created_at'] = reservation['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                if reservation['updated_at']:
                    reservation['updated_at'] = reservation['updated_at'].strftime('%Y-%m-%d %H:%M:%S')
                
                return render_template('edit_reservation.html', reservation=reservation)
            else:
                flash('Rezervasyon bulunamadı!', 'error')
                
        except Error as e:
            flash(f'Rezervasyon getirilirken hata oluştu: {e}', 'error')
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    else:
        flash('Veritabanı bağlantısı kurulamadı!', 'error')
    
    return redirect(url_for('reservations_list'))

@app.route('/reservation/update/<int:reservation_id>', methods=['POST'])
def update_reservation(reservation_id):
    """Rezervasyonu güncelle"""
    # Form verilerini al
    name_surname = request.form.get('name_surname', '').strip()
    center = request.form.get('center', '')
    date = request.form.get('date', '')
    time = request.form.get('time', '')
    description = request.form.get('description', '').strip()

    # Temel validasyon
    if not all([name_surname, center, date, time]):
        flash('Lütfen tüm zorunlu alanları doldurunuz!', 'error')
        return redirect(url_for('edit_reservation', reservation_id=reservation_id))

    # Çakışma kontrolü (kendisi hariç)
    if check_reservation_conflict(center, date, time, reservation_id):
        alternatives = get_alternative_times(center, date, time)
        if alternatives:
            alt_text = ", ".join(alternatives)
            flash(f'Bu saat dilimi ({time}) için {center} merkezinde {date} tarihinde zaten rezervasyon var! Alternatif saatler: {alt_text}', 'error')
        else:
            flash(f'Bu saat dilimi ({time}) için {center} merkezinde {date} tarihinde zaten rezervasyon var ve alternatif saat bulunamadı!', 'error')
        
        return redirect(url_for('edit_reservation', reservation_id=reservation_id))

    # Rezervasyonu güncelle
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            query = """
            UPDATE reservations 
            SET name_surname = %s, center = %s, date = %s, time = %s, description = %s
            WHERE id = %s
            """
            cursor.execute(query, (name_surname, center, date, time, description, reservation_id))
            connection.commit()
            
            if cursor.rowcount > 0:
                flash(f'#{reservation_id} numaralı rezervasyon başarıyla güncellendi!', 'success')
            else:
                flash('Rezervasyon bulunamadı!', 'error')
                
        except Error as e:
            flash(f'Rezervasyon güncellenirken hata oluştu: {e}', 'error')
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    else:
        flash('Veritabanı bağlantısı kurulamadı!', 'error')

    return redirect(url_for('reservations_list'))

def get_time_availability(center, date):
    """Belirli bir merkez ve tarih için saat durumlarını getir"""
    time_slots = [
        "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00",
        "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00",
        "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00", "21:00-22:00"
    ]
    
    connection = get_db_connection()
    if not connection:
        return {slot: 'bos' for slot in time_slots}
    
    try:
        cursor = connection.cursor()
        query = """
        SELECT time FROM reservations 
        WHERE center = %s AND date = %s AND status IN ('onay', 'bekle')
        """
        cursor.execute(query, (center, date))
        occupied_times = [row[0] for row in cursor.fetchall()]
        
        availability = {}
        for slot in time_slots:
            availability[slot] = 'dolu' if slot in occupied_times else 'bos'
        
        return availability
        
    except Error as e:
        print(f"Saat durumu kontrolü hatası: {e}")
        return {slot: 'bos' for slot in time_slots}
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

@app.route('/availability')
def availability_view():
    """Saat durumu görüntüleme sayfası"""
    selected_center = request.args.get('center', '')
    selected_date = request.args.get('date', '')
    
    centers = ['Sefaköy Kültür Merkezi', 'Cennet Kültür Merkezi', 'Atakent Kültür Merkezi', 'Kemalpaşa Semt Konağı']
    availability_data = {}
    
    if selected_center and selected_date:
        availability_data = get_time_availability(selected_center, selected_date)
    
    return render_template('availability.html', 
                         centers=centers,
                         selected_center=selected_center,
                         selected_date=selected_date,
                         availability_data=availability_data)

if __name__ == '__main__':
    # Uygulama başlatılırken veritabanını hazırla
    init_database()
    app.run(debug=True)